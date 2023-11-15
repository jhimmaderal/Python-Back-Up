import pyautogui as pg
import time
from openpyxl import load_workbook as lw
import costChange as cc

def appPriceChange():
    file = "fpAppFile.xlsx"
    itemPrice = []
    xslFile = lw(file)
    shtFile = xslFile["Price Change"] #change Sheet ["sheetName"]
    rowFile = shtFile.max_row  # count total row
    colFile = shtFile.max_column  # count total column
    
    supplier = input("Enter Supplier Code: ").upper()
    
    def goToPriceChange():
        pg.click(58,17, button='left')
        pg.keyDown('alt')
        pg.press('r',presses=2)
        pg.keyUp('alt')
        pg.press('enter')
        pg.press('p', presses= 2)
        pg.press('enter')   
        
    def createChange():
        pg.press("f9")
        pg.write("HO")
        pg.press('tab', presses= 5)
        pg.typewrite(supplier)
        pg.press('tab', presses= 13)
        pg.typewrite("PRICE CHANGE")
        
    def createLines():
        pg.press("f11")
        steps = 0
        for item in itemPrice:
            if steps == 0:
                if len(str(item)) == 6:
                    pg.write("@" + str(item))
                    steps = steps + 1
                else:
                    pg.write("@" + str(item))
                    steps = steps + 1
                
            else:
                pg.press("tab", presses=7)
                pg.write(".")
                pg.press("delete", presses=2)
                pg.press("left",presses=3)
                pg.write(str(item))
                pg.press("enter")
                time.sleep(0.5)
                steps = steps - 1                   
    
    def runApp():
        
        print("Choose Item: ")
        print("[1] Change Cost")
        print("[2] Price Change")
        selectApp = int(input("Select Item: "))
        startRow = 3
        
        if selectApp == "2": # Price Change
            print("Price Change - Running.....")   
            print("1. Create New")
            print("2. Continue Last Session")        
            createNew = input("Select Item: ")
            
            for item in range(startRow,rowFile + 1): # loop to A3 to last item
              codeCell = shtFile.cell(row=item, column = 1)
              codeVal = codeCell.value
              itemPrice.append(codeVal)
              retailCell = shtFile.cell(row=item, column = 4)
              retailVal = retailCell.value
              itemPrice.append(retailVal)

            try:
                if createNew == 1:
                    startRow = 3
                    pg.click(58,17, button='left')
                    time.sleep(2)
                    goToPriceChange()
                    createChange()

                    createLines()
                    print(itemPrice)
                    exit()
                elif createNew == 2:
                    startRow = input("Select Row:  Default is 3: ")
                    pg.click(58,17, button='left')
                    time.sleep(2)

                    createLines()
                    print(itemPrice)
                    exit()
                else:
                    print("Not in the list! ")
                    exit
            except ValueError:
                createNew = input("Select Item: ")
            
        elif selectApp == "1": # Cost Change
            cc.runCostChange()
            exit()
    runApp()
    
    
    
    
    